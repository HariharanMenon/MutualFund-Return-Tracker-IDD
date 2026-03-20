"""
conftest.py — Pytest fixtures shared across all backend test modules.

Scope strategy:
  - "session"  → expensive objects created once (TestClient, xlsx bytes)
  - "function" → stateful or mutation-sensitive fixtures
"""

import io

import openpyxl
import pytest
from starlette.testclient import TestClient

from main import app

# ---------------------------------------------------------------------------
# Internal helpers (not pytest fixtures)
# ---------------------------------------------------------------------------

_HEADERS = ["Date", "Transaction Type", "Amount", "Units", "Price", "Unit Balance"]


def make_xlsx_bytes(headers: list, rows: list) -> bytes:
    """Build an in-memory .xlsx file and return its bytes.

    Parameters
    ----------
    headers:
        Column header row (list of strings).
    rows:
        Data rows (list of lists). Use ``None`` for empty cells.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTTP test client
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def client() -> TestClient:
    """Starlette TestClient wrapping the FastAPI app.
    Remains alive for the full test session (stateless app, safe to reuse).
    """
    with TestClient(app) as c:
        yield c


# ---------------------------------------------------------------------------
# Xlsx byte fixtures — built once per session
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def valid_xlsx() -> bytes:
    """Minimal valid fund statement: Purchase → SELL."""
    return make_xlsx_bytes(
        _HEADERS,
        [
            ["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jan-2021", "SELL", 11500, 100.0, None, None],
        ],
    )


@pytest.fixture(scope="session")
def stamp_duty_xlsx() -> bytes:
    """Valid statement including a Stamp Duty row."""
    return make_xlsx_bytes(
        _HEADERS,
        [
            ["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jan-2020", "Stamp Duty", 50, None, None, None],
            ["01-Jan-2021", "SELL", 11500, 100.0, None, None],
        ],
    )


@pytest.fixture(scope="session")
def dividend_reinvest_xlsx() -> bytes:
    """Valid statement with Dividend Reinvest transaction."""
    return make_xlsx_bytes(
        _HEADERS,
        [
            ["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jul-2020", "Dividend Reinvest", 500, 5.0, 100.0, 105.0],
            ["01-Jan-2021", "SELL", 12000, 105.0, None, None],
        ],
    )


@pytest.fixture(scope="session")
def missing_date_col_xlsx() -> bytes:
    """Xlsx missing the Date column header."""
    return make_xlsx_bytes(
        ["Transaction Type", "Amount", "Units", "Price", "Unit Balance"],
        [
            ["Purchase", 10000, 100.0, 100.0, 100.0],
            ["SELL", 11500, 100.0, None, None],
        ],
    )


@pytest.fixture(scope="session")
def invalid_dates_xlsx() -> bytes:
    """Xlsx with a date in wrong format (DD/MM/YYYY instead of DD-MMM-YYYY)."""
    return make_xlsx_bytes(
        _HEADERS,
        [
            ["01/01/2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jan-2021", "SELL", 11500, 100.0, None, None],
        ],
    )


@pytest.fixture(scope="session")
def no_redemption_xlsx() -> bytes:
    """Xlsx where the last transaction is a Purchase (not SELL/REDEMPTION)."""
    return make_xlsx_bytes(
        _HEADERS,
        [
            ["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jun-2020", "Purchase", 5000, 45.0, 111.11, 145.0],
        ],
    )


@pytest.fixture(scope="session")
def multi_sheet_xlsx() -> bytes:
    """Xlsx with two worksheets — only the first should be read."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Transactions"
    ws1.append(_HEADERS)
    ws1.append(["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0])
    ws1.append(["01-Jan-2021", "SELL", 11500, 100.0, None, None])
    ws2 = wb.create_sheet("IgnoredSheet")
    ws2.append(["garbage", "data", "that", "should", "be", "ignored"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(scope="session")
def uppercase_headers_xlsx() -> bytes:
    """Xlsx with column headers in uppercase — must be accepted (case-insensitive)."""
    return make_xlsx_bytes(
        ["DATE", "TRANSACTION TYPE", "AMOUNT", "UNITS", "PRICE", "UNIT BALANCE"],
        [
            ["01-Jan-2020", "Purchase", 10000, 100.0, 100.0, 100.0],
            ["01-Jan-2021", "SELL", 11500, 100.0, None, None],
        ],
    )
